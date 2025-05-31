from modelos import Usuario, CuentaBancaria
import pandas as pd
import io
from openpyxl.styles import Font, Alignment
from datetime import datetime

class ExportacionServicio:
    def __init__(self, transaccion_repositorio, cuenta_bancaria_servicio, categoria_servicio, transaccion_servicio):
        self.transaccion_repositorio = transaccion_repositorio
        self.cuenta_bancaria_servicio = cuenta_bancaria_servicio
        self.categoria_servicio = categoria_servicio
        self.transaccion_servicio = transaccion_servicio

    def exportar_excel(self, usuario_id, cuenta_id):
        usuario = self.transaccion_repositorio.obtener_por_id(Usuario, usuario_id)
        if not usuario:
            raise LookupError("Usuario no encontrado")

        cuenta = self.transaccion_repositorio.obtener_unico_con_filtro(
            CuentaBancaria, [
                CuentaBancaria.id == cuenta_id,
                CuentaBancaria.usuario_id == usuario_id
            ]
        )
        if not cuenta:
            raise LookupError("Cuenta no encontrada")

        transacciones = self.transaccion_repositorio.obtener_transacciones_con_categoria(cuenta_id)

        data = [
            {
                "categoria": t.categoria.nombre,
                "tipo": "Ingreso" if t.categoria.tipo.name == "INGRESO" else "Gasto",
                "descripcion": t.descripcion,
                "monto": t.monto,
                "fecha": t.fecha
            }
            for t in transacciones
        ]
        df = pd.DataFrame(data, columns=["categoria", "tipo", "descripcion", "monto", "fecha"])

        total_ingresos = df[df['monto'] > 0]['monto'].sum() if not df.empty else 0
        total_gastos = abs(df[df['monto'] < 0]['monto'].sum()) if not df.empty else 0
        saldo_cuenta = cuenta.saldo
        usuario_nombre = usuario.nombre
        nombre_cuenta = cuenta.nombre

        df['Total'] = df['monto'].cumsum() + (saldo_cuenta - df['monto'].sum() if not df.empty else saldo_cuenta)
        df = df[['categoria', 'tipo', 'descripcion', 'monto', 'fecha', 'Total']]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Transacciones')
            ws = writer.sheets['Transacciones']

            # Formato columnas
            column_widths = {
                'A': 20, 'B': 15, 'C': 30, 'D': 15, 'E': 20, 'F': 15
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            for col in ['D', 'F']:
                for row in range(2, len(df) + 2):
                    cell = ws[f'{col}{row}']
                    cell.number_format = '"₡"#,##0.00'
            for row in range(2, len(df) + 2):
                monto_cell = ws[f'D{row}']
                if monto_cell.value and monto_cell.value < 0:
                    monto_cell.font = Font(color="FF0000")
            ws.insert_rows(1, 4)
            ws['A1'] = "Usuario:"; ws['B1'] = usuario_nombre
            ws['A2'] = "Cuenta:"; ws['B2'] = nombre_cuenta
            ws['A3'] = "Saldo Actual:"; ws['B3'] = saldo_cuenta
            ws['B3'].number_format = '"₡"#,##0.00'
            ws['A4'] = "Total Ingresos:"; ws['B4'] = total_ingresos
            ws['B4'].number_format = '"₡"#,##0.00'
            ws['D4'] = "Total Gastos:"; ws['E4'] = total_gastos
            ws['E4'].number_format = '"₡"#,##0.00'
            ws['E4'].font = Font(color="FF0000")
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            ws.freeze_panes = 'A6'

        output.seek(0)
        return output, usuario_nombre, nombre_cuenta

    def importar_excel(self, ruta_excel, usuario_id, cuenta_id):
        # 1. Buscar la fila que contiene los encabezados
        columnas_objetivo = ["categoria", "tipo", "descripcion", "monto", "fecha"]

        # Leer el archivo completo sin encabezado definido
        df_raw = pd.read_excel(ruta_excel, header=None)
        encabezado_idx = None

        for idx, row in df_raw.iterrows():
            # Normaliza y filtra valores no nulos
            row_norm = [str(c).strip().lower() for c in row.values if pd.notna(c)]
            if all(col in row_norm for col in columnas_objetivo):
                encabezado_idx = idx
                break

        if encabezado_idx is None:
            raise ValueError("No se encontraron los encabezados mínimos requeridos: categoria, tipo, descripcion, monto, fecha")

        # 2. Leer el archivo de nuevo, usando esa fila como encabezado
        df = pd.read_excel(ruta_excel, header=encabezado_idx)
        df.columns = [col.strip().lower() for col in df.columns]
        print("Columnas detectadas:", df.columns)

        # Filtra solo las filas que tengan 'categoria' no vacía
        df = df.dropna(subset=["categoria"])
        print("DataFrame shape:", df.shape)
        print(df.head(10))

        # Normaliza tipos de texto
        for columna in ["categoria", "tipo", "descripcion"]:
            df[columna] = df[columna].astype(str)

        errores = []
        importadas = 0

        # Validaciones previas de IDs
        try:
            cuenta_id = int(cuenta_id)
            usuario_id = int(usuario_id)
        except (ValueError, TypeError):
            raise ValueError("ID de cuenta o usuario inválido.")

        cuenta = self.cuenta_bancaria_servicio.obtener_cuenta_por_id(cuenta_id)
        if not cuenta:
            raise ValueError("Cuenta no encontrada.")
        if int(cuenta.usuario_id) != usuario_id:
            raise ValueError("La cuenta no pertenece al usuario autenticado.")

        categorias_actuales = self.categoria_servicio.obtener_categorias(cuenta_id)
        categorias_map = {cat.nombre.strip(): cat for cat in categorias_actuales}

        for idx, fila in df.iterrows():
            try:
                nombre_categoria = str(fila["categoria"]).strip()
                tipo = str(fila["tipo"]).strip().upper()
                descripcion = str(fila["descripcion"]).strip()

                try:
                    monto = abs(float(fila["monto"]))
                except (ValueError, TypeError):
                    errores.append(f"Fila {idx+encabezado_idx+2}: El monto no es un número válido.")
                    continue

                try:
                    fecha = pd.to_datetime(fila["fecha"])
                except Exception:
                    errores.append(f"Fila {idx+encabezado_idx+2}: La fecha no tiene un formato válido.")
                    continue

                categoria = categorias_map.get(nombre_categoria)
                if not categoria:
                    try:
                        categoria = self.categoria_servicio.crear_categoria(
                            nombre=nombre_categoria,
                            tipo=tipo,
                            presupuesto=None,
                            cuenta_id=cuenta_id
                        )
                        categorias_map[nombre_categoria] = categoria
                    except Exception as e:
                        raise ValueError(f"Fila {idx+encabezado_idx+2}: Error al crear la categoría '{nombre_categoria}': {str(e)}")

                    if categoria is None:
                        # Maneja el error, categoría inexistente
                        raise ValueError("Categoría no encontrada.")
                    
                ya_existe = self.transaccion_servicio.transaccion_duplicada(
                    cuenta_id=cuenta_id,
                    categoria=categoria,  # <-- Aquí ahora pasas el objeto, que ya tienes arriba
                    descripcion=descripcion,
                    monto=monto,
                    fecha=fecha
                )
                if ya_existe:
                    errores.append(f"Fila {idx+5}: Duplicado detectado (no se importa de nuevo)")
                    continue


                self.transaccion_servicio.registrar_transaccion(
                    cuenta_id=cuenta_id,
                    categoria_id=categoria.id,
                    descripcion=descripcion,
                    monto=monto,
                    fecha=fecha,
                )
                importadas += 1
            except Exception as e:
                print(f"Error en fila {idx}: {e}")
                errores.append(f"Fila {idx+encabezado_idx+2}: {str(e)}")

        return {"ok": len(errores) == 0, "importadas": importadas, "errores": errores}
